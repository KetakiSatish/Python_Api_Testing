import pytest
import requests
import json
import jsonpath
import openpyxl


class common:
    def __init__(self, filenamepath, sheetname):
        global wb
        global sh
        wb = openpyxl.load_workbook(filenamepath)
        sh = wb[sheetname]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        columns = sh.max_column
        return columns

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self,rownum,jsonrequest,keylist):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rownum, column=i)
            jsonrequest[keylist[i-1]] = cell.value

        return jsonrequest

