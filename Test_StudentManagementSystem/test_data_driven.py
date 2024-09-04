import requests
import json
import openpyxl
import jsonpath


def test_data_driven():
    url = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("C:\\Users\\Ketaki\\Desktop\\API\\data drivn.json", 'r')
    wb = openpyxl.load_workbook('C:\\Users\\Ketaki\\Desktop\\API\\data_driven.xlsx')
    sh = wb['Sheet1']
    rows = sh.max_row
    json_req = json.loads(f.read())
    for i in range(2, rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)
        json_req['first_name'] = cell_first_name.value
        json_req['middle_name'] = cell_middle_name.value
        json_req['last_name'] = cell_last_name.value
        json_req['date_of_birth'] = cell_date_of_birth.value
    resp = requests.post(url, json_req)
    print(resp.status_code)
    assert resp.status_code == 201
